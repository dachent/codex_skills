#!/usr/bin/env python3
"""Scan OOXML Excel workbooks for formula cells and visible Excel error cells."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
EXCEL_ERRORS = [
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
    "#SPILL!",
    "#CALC!",
]
MAX_LOCATIONS_PER_ERROR = 50


def workbook_ref(path: Path) -> str:
    return str(path.resolve(strict=False))


def build_result(
    *,
    status: str,
    path: Path,
    message: str | None = None,
    error_kind: str | None = None,
    total_formulas: int = 0,
    total_errors: int = 0,
    error_summary: dict | None = None,
) -> dict:
    result = {
        "status": status,
        "workbook": workbook_ref(path),
        "total_formulas": total_formulas,
        "total_errors": total_errors,
        "error_summary": error_summary or {},
    }
    if message is not None:
        result["message"] = message
    if error_kind is not None:
        result["error_kind"] = error_kind
    if status == "error":
        result["supported_extensions"] = sorted(SUPPORTED_EXTENSIONS)
    return result


def error_result(path: Path, message: str, error_kind: str) -> dict:
    return build_result(status="error", path=path, message=message, error_kind=error_kind)


def scan_workbook(path: Path) -> dict:
    suffix = path.suffix.lower()
    if not path.exists():
        return error_result(path, f"File does not exist: {path}", "missing_file")

    if suffix not in SUPPORTED_EXTENSIONS:
        return error_result(
            path,
            (
                f"Unsupported workbook format '{suffix or '<none>'}'. "
                "Validate only OOXML Excel workbooks after converting to "
                ".xlsx, .xlsm, .xltx, or .xltm."
            ),
            "unsupported_extension",
        )

    wb_values = None
    wb_formulas = None
    try:
        wb_values = load_workbook(path, data_only=True)
        wb_formulas = load_workbook(path, data_only=False)
    except (InvalidFileException, BadZipFile, EOFError, KeyError, OSError, ValueError) as exc:
        return error_result(path, f"Failed to load workbook: {exc}", "load_failed")
    except Exception as exc:
        return error_result(path, f"Unexpected workbook load failure: {exc}", "load_failed")

    error_details = {err: [] for err in EXCEL_ERRORS}
    error_counts = {err: 0 for err in EXCEL_ERRORS}
    total_errors = 0
    total_formulas = 0

    try:
        for sheet_name in wb_formulas.sheetnames:
            ws_values = wb_values[sheet_name]
            ws_formulas = wb_formulas[sheet_name]

            for row_values, row_formulas in zip(ws_values.iter_rows(), ws_formulas.iter_rows()):
                for cell_value, cell_formula in zip(row_values, row_formulas):
                    if cell_formula.data_type == "f":
                        total_formulas += 1

                    if cell_value.data_type != "e":
                        continue

                    error_value = cell_value.value
                    if error_value not in error_details:
                        continue

                    error_counts[error_value] += 1
                    locations = error_details[error_value]
                    if len(locations) < MAX_LOCATIONS_PER_ERROR:
                        locations.append(f"{sheet_name}!{cell_value.coordinate}")
                    total_errors += 1
    finally:
        wb_values.close()
        wb_formulas.close()

    error_summary = {
        err: {
            "count": error_counts[err],
            "locations": locations,
        }
        for err, locations in error_details.items()
        if error_counts[err]
    }

    if total_errors == 0:
        return build_result(
            status="success",
            path=path,
            message="No visible Excel error cells were found.",
            total_formulas=total_formulas,
            total_errors=0,
            error_summary={},
        )

    return build_result(
        status="errors_found",
        path=path,
        message="Visible Excel error cells were found.",
        total_formulas=total_formulas,
        total_errors=total_errors,
        error_summary=error_summary,
    )


def emit_json(result: dict) -> int:
    print(json.dumps(result, indent=2))
    status = result.get("status")
    if status == "success":
        return 0
    if status == "errors_found":
        return 2
    return 1


def main() -> int:
    if len(sys.argv) != 2:
        usage_result = {
            "status": "error",
            "message": "Usage: python check_formula_errors.py <workbook_path>",
            "error_kind": "invalid_arguments",
            "supported_extensions": sorted(SUPPORTED_EXTENSIONS),
        }
        return emit_json(usage_result)

    return emit_json(scan_workbook(Path(sys.argv[1])))


if __name__ == "__main__":
    raise SystemExit(main())
